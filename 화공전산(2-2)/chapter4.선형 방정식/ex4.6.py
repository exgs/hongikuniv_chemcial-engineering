import math
import openpyxl as xl
import os

# a = os.getcwd()
# os.chdir("C:\\Users\\이윤성\\Desktop\\python\\hongik")
# os.chdir(".\\hongik")
wb = xl.load_workbook(filename='chapter4_example.xlsx', data_only=True)
ws = wb['1']
ra, rb = 2,5
ca, cb, cc = 1, 4, 5
A = [[ws.cell(i,j).value for j in range(ca, cb+1)] for i in range(ra, rb + 1)]
b = [ws.cell(i,cc).value for i in range(ra, rb+1)]
print(A)
print(b)
wb.close()

import numpy.linalg
import scipy.linalg

x1 = numpy.linalg.solve(A,b)
x2 = scipy.linalg.solve(A,b)
print(x1,x2)
