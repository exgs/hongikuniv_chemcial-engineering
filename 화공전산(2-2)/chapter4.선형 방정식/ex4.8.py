import numpy
import scipy.linalg
import openpyxl as xl
from numpy import transpose, matmul
import os

# os.chdir(".\\hongik")
wb = xl.load_workbook("chapter4_example.xlsx")
ws = wb['1']
ra, rb = 2,5
ca, cb, cc = 1, 4, 5
A = [[ws.cell(i,j).value for j in range(ca, cb+1)] for i in range(ra, rb + 1)]
b = [ws.cell(i,cc).value for i in range(ra, rb+1)]
# print(A)
# print(b)
wb.close()

P,L,U = scipy.linalg.lu(A)
print(P)
print(L)
print(U)
Pt = transpose(P)
z = matmul(Pt,b)
y = scipy.linalg.solve_triangular(L,z, lower = True)
x = scipy.linalg.solve_triangular(U,y, lower = False)

print("-----------------answer--------------------")
print(x)