import scipy
import numpy
import openpyxl as xl

wb = xl.load_workbook("p4.xlsx")
ws = wb['p4.2']
ra, rb=2,5
ca, cb=1,4
A = [[ws.cell(i,j).value for j in range(ca,cb+1)] for i in range(ra,rb+1)]
b = [ws.cell(i,cb+1).value for i in range(ra,rb+1)]
print(A)
print(b)
wb.close()
# from scipy.linalg import solve
from numpy.linalg import solve, det
print(numpy.linalg.det(A)) #det = 0
x = numpy.linalg.solve(A,b)
print(x)