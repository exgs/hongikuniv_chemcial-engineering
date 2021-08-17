import openpyxl as xl
import scipy.linalg

wb = xl.load_workbook("p4.xlsx")
ws = wb['pr0404']
ra,rb = 2,8
ca,cb = 1,7
A = [[ws.cell(i,j).value for j in range(ca,cb+1)] for i in range(ra,rb+1)]
b = [ws.cell(i,cb+1).value for i in range(ra,rb+1)]
# print(scipy.linalg.det(A))
print(A,b)
wb.close()
x = scipy.linalg.solve(A,b)
for i in range(len(x)):
	print("x[%d] = %8.4f" %(i,x[i]))