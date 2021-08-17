import scipy
import numpy
import openpyxl as xl

wb = xl.load_workbook("p4.xlsx")
ws = wb['p4.1']
ra, rb=2,4
ca, cb=1,3
A = [[ws.cell(i,j).value for j in range(ca,cb+1)] for i in range(ra,rb+1)]
b = [ws.cell(i,cb+1).value for i in range(ra,rb+1)]
print(A)
print(b)
wb.close()
from scipy.linalg import solve, det
print(scipy.linalg.det(A)) #det = 0
x = scipy.linalg.solve(A,b)
print(x)