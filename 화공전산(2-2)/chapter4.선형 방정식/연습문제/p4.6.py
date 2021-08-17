import openpyxl as xl
import scipy.linalg

wb = xl.load_workbook("p4.xlsx")
ws = wb['sparse']
ra,rb = 1,20
ca,cb = 1,20
A = [[ws.cell(i,j).value for j in range(ca,cb+1)] for i in range(ra,rb+1)]
b = [ws.cell(i,cb+1).value for i in range(ra,rb+1)]
for i in range(20):
	for j in range(20):
		if (A[i][j] == None):
			A[i][j] = 0
# print(A,b)
print(scipy.linalg.det(A))
wb.close()

import numpy.linalg
x = numpy.linalg.solve(A,b)
k = numpy.linalg.cond(A)
print("k : %10.1f" %k)
for i in range(len(x)):
	print("x[%d] = %8.4f" %(i,x[i]))