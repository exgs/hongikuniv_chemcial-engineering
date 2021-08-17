import openpyxl as xl
wb = xl.load_workbook(filename = './chapter11.최소 제곱 곡선 맞춤/curvefit.xlsx', data_only = True, keep_vba = True)
ws = wb['vp']
ra, rb = 2, 19
n = rb - ra + 1
Td = [ws.cell(r,1).value for r in range(ra, rb+1)]
pd = [ws.cell(r,2).value for r in range(ra, rb+1)]
wb.close()

from math import log, exp

def f_absolute(T, a, b, c):
	sol = exp(a - b/(T-c))
	return sol
def f_relative(T, a, b, c):
	sol = (a - b/(T-c))
	return sol

def fun(coef, Td, pd):
	evec = []
	for i in range(0, n):
		ei = coef[0] - coef[1] / (Td[i] - coef[2]) - log(pd[i])
		evec.append(ei)
	return evec

coef = [0, 0, 0]
from scipy.optimize import least_squares
r = least_squares(fun, coef, args = (Td, pd))
a = r.x[0]; b = r.x[1]; c = r.x[2]

import numpy
tj1 = numpy.linspace(50, 400, 51)
from matplotlib import pyplot as plt
# plt.plot(Td, list(map(lambda x : log(x), pd)))
# plt.plot(tj1, [f_relative(i,a,b,c) for i in tj1], '+')
# plt.yscale('log')
# plt.show()

fig, ax = plt.subplots(2)
ax[0].plot(Td, pd)
ax[0].plot(tj1, [f_absolute(i,a,b,c) for i in tj1], '+')

ax[1].plot(Td, list(map(lambda x : log(x), pd)))
ax[1].plot(tj1, [f_relative(i,a,b,c) for i in tj1], '+')
ax[1].set_yscale('log')
plt.show()