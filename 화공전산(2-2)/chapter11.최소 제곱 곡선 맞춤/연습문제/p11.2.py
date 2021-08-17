import openpyxl as xl
wb = xl.load_workbook(filename = './p11.xlsx', data_only = True)
ws = wb['p11.2']
ra, rb = 4, 11
n = rb - ra + 1
Td = [ws.cell(r,2).value for r in range(ra, rb+1)]
pd = [ws.cell(r,3).value for r in range(ra, rb+1)]
wb.close()

from math import log, exp

def fun(coef, Td, kr):
	evec = []
	R = 8.314
	for i in range(0, n):
		# ei = coef[0] * coef[1] / (Td[i] - coef[2]) - log(kr[i])
		ei = coef[0] * exp(-coef[1]/(R*Td[i])) - kr[i]
		evec.append(ei)
	return evec
print(Td)
print(pd)
print('-'*30)
coef = [0, 0]
from scipy.optimize import least_squares
r = least_squares(fun, coef, args = (Td, pd))
print(r.x)