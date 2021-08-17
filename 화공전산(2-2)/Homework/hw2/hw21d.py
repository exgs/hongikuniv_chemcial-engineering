import openpyxl as xl

wb = xl.load_workbook("B419111hw20.xlsx", data_only = True)
ws = wb['hw21']
ra, rb = 3, 12
ca, cb = 2, 11
A = [[ws.cell(i,j).value for j in range(ca,cb+1)] for i in range(ra, rb + 1)]
b = [ws.cell(i, cb + 1).value for i in range(ra, rb+1)]
wb.close()

wb = xl.load_workbook("B419111hw20.xlsx", data_only = False)
ws = wb['hw21']

import numpy
x = numpy.linalg.solve(A,b)
x_c = cb + 5 # P열
j = 0
for i in range(ra, rb+1):
	ws.cell(i, x_c).value = x[j]
	j = j+1
wb.save("B419111hw20.xlsx")
wb.close()
input("Press Enter")