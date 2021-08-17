import openpyxl as xl
wb = xl.load_workbook(filename='../스플라인보간법.xlsx', data_only = True)
ws = wb['표10.5']
ra, rb = 3, 12
jp = 2
jT = 4 #jT= 3 for iso-pentane, 4 for ammonia, 5 for diethyl ether
pd = [ws.cell(r,jp).value for r in range(ra, rb+1)]
name = ws.cell(2,jT).value
Td = [ws.cell(r,jT).value for r in range(ra, rb+1)]
wb.close()
# n = rb - ra + 1
n = rb - ra
import numpy
coef = numpy.zeros((4, n))

def f(b0):
	i = 0
	hi = Td[i + 1] - Td[i]
	Di = (pd[i +1] - pd[i]) / hi
	coef[0][i] = pd[i]
	coef[1][i] = b0
	coef[2][i] = 0
	coef[3][i] = (Di - coef[1][i] - coef[2][i] * hi) / hi**2
	while i < n-1:
		j, i = i, i + 1	# j = i-1
		hj, hi = hi, Td[i+1] - Td[i] # hj = h(i-1)
		Di = (pd[i+1] - pd[i]) / hi
		coef[0][i] = pd[i]
		coef[1][i] = coef[2][j] + 3*hj*coef[3][j]
		coef[2][i] = coef[1][j] + 2*coef[2][j]*hj + 3*coef[3][j]*hj**2
		coef[3][i] = (Di - coef[1][i] - coef[2][i] * hi) / hi**2
	return coef[1][j] + 2*coef[2][j]*hj + 3*coef[3][j]*hj**2
	
from scipy.optimize import bisect
bisect(f, 0,1)
print(coef)
print("Nautral spline for " + name)
print("%10s%10s%10s%10s" %('a','b','c','d'))
for j in range(0, n):
	for i in range(0,4):
		print("%10.5f" %coef[i,j], end = " ")
	print()