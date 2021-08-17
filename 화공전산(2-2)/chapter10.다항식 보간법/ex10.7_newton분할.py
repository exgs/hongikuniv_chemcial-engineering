import openpyxl as xl
wb = xl.load_workbook(filename='.\chapter10.다항식 보간법\스플라인보간법.xlsx', data_only = True)
ws = wb['표10.5']
ra, rb = 3, 12
jp = 2
jT = 4 #jT= 3 for iso-pentane, 4 for ammonia, 5 for diethyl ether
pd = [ws.cell(r,jp).value for r in range(ra, rb+1)]
name = ws.cell(2,jT).value
Td = [ws.cell(r,jT).value for r in range(ra, rb+1)]
wb.close()
n = rb - ra + 1

import numpy
coef = numpy.zeros((n, n))
#여기서 p값이 y값이고 T값이 x값이다!!
for i in range(0, n):
	coef[i,0] = pd[i]
for j in range(1, n): # newton's divided difference
	for i in range(0, n - j):
		num = coef[i + 1, j - 1] - coef[i, j -1]
		den = Td[i + j] - Td[i]
		coef[i,j] = num/den

def p(x, xd, coef, n): #interpolating polynomial
	y, xx = 0, 1
	for j in range(0, n):
		y = y + coef[0, j] * xx
		xx = xx * (x - xd[j])
	return y

Ts = numpy.linspace(Td[0], Td[-1], 128)
from matplotlib import pyplot as plt
plt.figure(figsize=(4,3))
plt.title(name)
plt.plot(Ts, [p(T, Td, coef, n) for T in Ts])
plt.plot(Td, pd, 'ko', mfc = 'w')
plt.tight_layout()
plt.show()